import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from io import BytesIO

# Streamlit UI setup
st.set_page_config(page_title="Excel Data Cleaner", page_icon="📊", layout="wide")
st.logo("logo.png", size = "large")
st.title('Excel Data Cleaner for MCHX (Charge-EEV) Experimental Results')
st.markdown("""
This tool automates the cleanup of experimental data scattered across multiple Excel sheets. 
Upload your Excel file below to consolidate all data into a clean, formatted table.
""")

# File uploader
uploaded_file = st.file_uploader('Upload your Excel file (.xlsx)', type=['xlsx'],
                                 help="File should contain multiple sheets with identical formatting")

if uploaded_file is not None:
    try:
        # Load workbook from uploaded file
        wb = load_workbook(uploaded_file, data_only=True)
        records = []

        # Display processing status
        with st.status("Processing your file...", expanded=True) as status:
            st.write("📂 Reading uploaded workbook...")

            # Define cell map (1-based indices)
            cell_map = {
                'Refrigerant Qty (g)': (16, 2),  # Converted to grams without suffix
                'EEV': (7, 2),
                'Test Condition': (2, 2),
                'Comp RPM/Hz': (4, 2),
                'IDU RPM': (5, 2),
                'ODU RPM': (6, 2),
                'ID DBT': (9, 2),
                'ID WBT': (10, 2),
                'OD DBT': (11, 2),
                'OD WBT': (12, 2),
                'Suction Pressure (Bar)': (19, 2),
                'Cond In Pressure (Bar)': (20, 2),
                'Cond Out Pressure (Bar)': (21, 2),
                'Suction Sat T (°C)': (19, 5),
                'Cond in Sat T (°C)': (20, 5),
                'Cond out Sat T (°C)': (21, 5),
                'Temp Compressor in (°C)': (23, 2),
                'Temp Condenser in (°C)': (24, 2),
                'Temp Condenser out (°C)': (25, 2),
                'Temp Evaporator out (°C)': (26, 2),
                'SH (Evaporator) (°C)': (23, 5),
                'SH (Suction) (°C)': (24, 5),
                'SC (°C)': (25, 5),
                'Power Consumption (W)': (29, 2),
                'Enthalpy Evap In (kJ/kg)': (31, 2),
                'Enthalpy Evap Out (kJ/kg)': (32, 2),
                '∆H Evap (kJ/kg)': (33, 2),
                'Enthalpy Cond In (kJ/kg)': (34, 2),
                'Cond Enthalpy Out (kJ/kg)': (35, 2),
                '∆H Cond (kJ/kg)': (36, 2),
                'Mass Flow (kg/hr)': (38, 2),
                'Cooling Cap (W)': (40, 2),
                'Cooling Cap (Btu/hr)': (40, 4),
                'COP (W/W)': (45, 2)
            }

            # Extract data from each worksheet
            sheet_count = len(wb.worksheets)
            st.write(f"🔍 Found {sheet_count} worksheets. Extracting data...")

            progress_bar = st.progress(0)
            for i, ws in enumerate(wb.worksheets):
                row_data = {}
                for header, (r, c) in cell_map.items():
                    cell_value = ws.cell(row=r, column=c).value

                    # Special handling for refrigerant quantity
                    if header == 'Refrigerant Qty (g)':
                        if isinstance(cell_value, str):
                            row_data[header] = int(cell_value.removesuffix('GM'))
                        else:
                            row_data[header] = cell_value
                    else:
                        row_data[header] = cell_value

                records.append(row_data)
                progress_bar.progress((i + 1) / sheet_count)

            # Build DataFrame
            df = pd.DataFrame(records)
            st.write(f"✅ Extracted {len(df)} records from {sheet_count} sheets")

            # Create new workbook
            out_wb = Workbook()
            ws_out = out_wb.active
            ws_out.title = 'Consolidated Data'

            # Write headers
            for col_idx, column_name in enumerate(df.columns, start=1):
                ws_out.cell(row=1, column=col_idx, value=column_name)

            # Write data rows
            st.write("📝 Creating cleaned workbook...")
            for row_idx, record in enumerate(df.itertuples(index=False), start=2):
                for col_idx, cell_value in enumerate(record, start=1):
                    ws_out.cell(row=row_idx, column=col_idx, value=cell_value)

            # Save to BytesIO buffer
            output = BytesIO()
            out_wb.save(output)
            output.seek(0)

            status.update(label="Processing complete! ✅", state="complete", expanded=False)

        # Show preview
        st.subheader("Preview of Cleaned Data")
        st.dataframe(df.head(), use_container_width=True)

        # Download button
        st.download_button(
            label='📥 Download Cleaned MCHX Excel File',
            data=output,
            file_name='cleaned_experimental_data.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            help="Click to download the consolidated data in Excel format"
        )

    except Exception as e:
        st.error(f"❌ Error processing file: {str(e)}")
        st.info("Please ensure your file matches the expected format and try again.")
else:
    st.info(
        "ℹ️ Please upload an Excel file to get started. Your file should have multiple sheets with identical formatting.")
